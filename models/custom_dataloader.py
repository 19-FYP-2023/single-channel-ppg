import os
import torch
import numpy as np 
import matplotlib.pyplot as plt 
from torch.utils.data import Dataset, DataLoader 
from openpyxl import load_workbook

class PpgBpDataset(Dataset):
    def __init__(self, root_dir, data_info_file, data_dir, samples_per_subject, transform=None):
        self.root_dir = root_dir
        self.data_info_file = data_info_file
        self.data_dir = data_dir
        self.transform = transform 
        self.workbook = load_workbook(f"{self.root_dir}/{self.data_info_file}")
        self.info_sheet = self.workbook["cardiovascular dataset"]
        self.samples_per_subject = samples_per_subject
        

        header_row = self.info_sheet[2]
        header_row = [cell.value for cell in header_row]

        # print(header_row)


        self.num_col_idx = header_row.index("Num.") + 1
        self.subject_col_idx = header_row.index("subject_ID") + 1
        self.sysbp_col_idx = header_row.index("Systolic Blood Pressure(mmHg)") + 1
        self.diabp_col_idx = header_row.index("Diastolic Blood Pressure(mmHg)") + 1


        # print(num_col_idx, self.subject_col_idx, sysbp_col_idx, diabp_col_idx)

    def __len__(self):
        return self.info_sheet.cell(self.info_sheet.max_row, 0).value * self.samples_per_subject

    def __getitem__(self, idx):
        dataset_idx = int(idx/self.samples_per_subject)
        sample_num = idx%self.samples_per_subject + 1

        row_idx = 3 + dataset_idx
        subject_id = self.info_sheet.cell(row_idx, self.subject_col_idx).value
        sysbp = self.info_sheet.cell(row_idx, self.sysbp_col_idx).value
        diabp = self.info_sheet.cell(row_idx, self.diabp_col_idx).value

        data_sample_path = f"{self.root_dir}/{self.data_dir}/{subject_id}_{sample_num}.txt"

        data_sample = np.genfromtxt(data_sample_path, delimiter='\t')
        data_sample = data_sample[:-1]
        
        return data_sample, subject_id, sysbp, diabp


