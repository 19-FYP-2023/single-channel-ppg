import numpy as np
import pandas as pd
import h5py
from models.resample import resample
from torch.utils.data import Dataset
from openpyxl import load_workbook


class PpgBpDataset(Dataset):
    def __init__(
        self, root_dir, data_info_file, data_dir, samples_per_subject, transform=None
    ):
        self.root_dir = root_dir
        self.data_info_file = data_info_file
        self.data_dir = data_dir
        self.transform = transform
        self.workbook = load_workbook(f"{self.root_dir}/{self.data_info_file}")
        self.info_sheet = self.workbook["cardiovascular dataset"]
        self.samples_per_subject = samples_per_subject

        header_row = self.info_sheet[2]
        header_row = [cell.value for cell in header_row]

        # print(header_row)

        self.num_col_idx = header_row.index("Num.") + 1
        self.subject_col_idx = header_row.index("subject_ID") + 1
        self.sysbp_col_idx = header_row.index("Systolic Blood Pressure(mmHg)") + 1
        self.diabp_col_idx = header_row.index("Diastolic Blood Pressure(mmHg)") + 1

        # print(num_col_idx, self.subject_col_idx, sysbp_col_idx, diabp_col_idx)

    def __len__(self):
        return (
            self.info_sheet.cell(self.info_sheet.max_row, 1).value
            * self.samples_per_subject
        )

    def __getitem__(self, idx):
        dataset_idx = int(idx / self.samples_per_subject)
        sample_num = idx % self.samples_per_subject + 1

        row_idx = 3 + dataset_idx
        subject_id = self.info_sheet.cell(row_idx, self.subject_col_idx).value
        sysbp = self.info_sheet.cell(row_idx, self.sysbp_col_idx).value
        diabp = self.info_sheet.cell(row_idx, self.diabp_col_idx).value

        data_sample_path = (
            f"{self.root_dir}/{self.data_dir}/{subject_id}_{sample_num}.txt"
        )

        data_sample = np.genfromtxt(data_sample_path, delimiter="\t")
        data_sample = data_sample[:-1]

        return data_sample, subject_id, sysbp, diabp


class ENTCDataset(Dataset):
    def __init__(self, root_dir, transforms=None):
        self.root_dir = root_dir
        self.label_path = self.root_dir + "good_signal_labels_with_paths.csv"
        self.label_csv = pd.read_csv(self.label_path)

    def __len__(self):
        return self.label_csv.shape[0]

    def __getitem__(self, idx):
        data_sample_path = self.root_dir + self.label_csv.iloc[idx]["Path"]
        data_sample_frame = pd.read_csv(data_sample_path)

        rows_to_drop = list(range(0, 8))
        data_sample_frame.drop(rows_to_drop, inplace=True)

        data_sample_frame[["VATA", "PITTA", "KAPHA"]] = data_sample_frame[
            list(data_sample_frame.columns)[0]
        ].apply(lambda x: pd.Series(str(x).strip().split(" ")))
        data_sample_frame["VATA"] = data_sample_frame["VATA"].apply(lambda x: int(x))
        data_sample_frame["PITTA"] = data_sample_frame["PITTA"].apply(lambda x: int(x))
        data_sample_frame["KAPHA"] = data_sample_frame["KAPHA"].apply(lambda x: int(x))

        data_sample_frame = data_sample_frame.reset_index()

        data_sample_frame.drop(
            list(data_sample_frame.columns)[:2], axis=1, inplace=True
        )

        data_sample = data_sample_frame.to_numpy()
        data_sample = data_sample.T
        return data_sample[:, :1024]


class UCIBPDataset(Dataset):
    def __init__(self, root_dir):
        self.root = root_dir
        self.count = []

        for i in range(0, 4):
            with h5py.File(f"{self.root}/Part_{i+1}.mat", "r") as f:
                self.count.append(len(f[f"Part_{i+1}"]))

    def __len__(self):
        return sum(self.count)

    def __getitem__(self, idx):
        partition_idx = idx // 3000
        sample_idx = idx % 3000

        with h5py.File(f"{self.root}/Part_{partition_idx+1}.mat", "r") as f:
            ori_data_sample_ppg = f[f[f"Part_{partition_idx+1}"][sample_idx][0]][:, 0]
            ori_data_sample_abp = f[f[f"Part_{partition_idx+1}"][sample_idx][0]][:, 1]

            abp_max = 200
            abp_min = 50
            ppg_max = 4.003
            ppg_min = 0.0

            original_freq = 125
            target_freq = 250

            upsample_factor = target_freq / original_freq

            data_sample_ppg = resample(
                data_sample_ppg[:1024], int(upsample_factor * 1024)
            ).astype(np.float32)
            data_sample_abp = resample(
                data_sample_abp[:1024], int(upsample_factor * 1024)
            ).astype(np.float32)

            return data_sample_ppg[:1024], data_sample_abp[:1024]


class UCIBPDatasetRaw(Dataset):
    def __init__(self, root_dir):
        self.root = root_dir
        self.count = []

        for i in range(0, 4):
            with h5py.File(f"{self.root}/Part_{i+1}.mat", "r") as f:
                self.count.append(len(f[f"Part_{i+1}"]))

    def __len__(self):
        return sum(self.count)

    def __getitem__(self, idx):
        partition_idx = idx // 3000
        sample_idx = idx % 3000

        with h5py.File(f"{self.root}/Part_{partition_idx+1}.mat", "r") as f:
            ori_data_sample_ppg = f[f[f"Part_{partition_idx+1}"][sample_idx][0]][:, 0]
            ori_data_sample_abp = f[f[f"Part_{partition_idx+1}"][sample_idx][0]][:, 1]

            abp_max = 200
            abp_min = 50
            ppg_max = 4.003
            ppg_min = 0.0

            original_freq = 125
            target_freq = 250

            upsample_factor = target_freq / original_freq
            data_sample_ppg = resample(
                ori_data_sample_ppg, int(upsample_factor * len(ori_data_sample_ppg))
            ).astype(np.float32)
            data_sample_abp = resample(
                ori_data_sample_abp, int(upsample_factor * len(ori_data_sample_abp))
            ).astype(np.float32)

            return (
                data_sample_ppg,
                data_sample_abp,
                ori_data_sample_ppg,
                ori_data_sample_abp,
            )


class UCIBPDatasetNPY(Dataset):
    def __init__(self, root_dir):
        self.root = root_dir
        self.abp_samples = np.load(f"{self.root}/abp_samples.npy")
        self.ppg_samples = np.load(f"{self.root}/ppg_samples.npy")
        self.pressures = np.load(f"{self.root}/pressures.npy")

    def __len__(self):
        return self.abp_samples.shape[0]

    def __getitem__(self, idx):
        return self.ppg_samples[idx], self.abp_samples[idx], self.pressures[idx]
